# app/routers/ops_dashboard.py
# -*- coding: utf-8 -*-
"""
Ops Security Dashboard (production-oriented)

Tệp này cung cấp "dashboard vận hành" (Ops) phục vụ giám sát bảo mật / an ninh theo thời gian thực.

Các endpoint chính:
- HTML:   GET  /ops/dashboard
          → trả file tĩnh app/static/ops_dashboard.html (UI realtime)
- SSE:    GET  /ops/sse
          → stream snapshot metrics định kỳ (mặc định 2.5s/lần)
- JSON:   GET  /ops/metrics/summary
          → timeseries 10 phút gần nhất (dùng fallback/Excel)
          GET  /ops/metrics/top_suspicious
          GET  /ops/metrics/current_bans
- Export: GET  /ops/export/metrics.xlsx
          → xuất Excel tổng hợp

Bảo vệ:
- Tất cả endpoint JSON/SSE/Export đều yêu cầu quyền Admin/Boss.
- SSE thường khó gắn header Authorization nếu dùng EventSource, vì vậy hỗ trợ nhiều nguồn token:
  (1) Query param ?token=...
  (2) Header Authorization: Bearer <token>   (khuyến nghị nếu client dùng fetch-stream)
  (3) Cookie 'access_token' (HttpOnly)       (tuỳ chọn)

Ghi chú quan trọng cho production:
- SSE là kết nối dài; nếu trong handler async bạn gọi Redis đồng bộ (sync) quá nặng, event-loop sẽ bị block.
  Vì vậy phần SSE trong file này offload các thao tác Redis/verify token (sync) sang thread bằng asyncio.to_thread().
- Nếu bạn đặt Nginx trước FastAPI, cần tắt buffering cho SSE (xem thêm header X-Accel-Buffering).
"""

from __future__ import annotations

# ===== Standard library =====
import asyncio  # ✅ cho sleep + to_thread trong SSE
import io       # ✅ tạo BytesIO để xuất Excel
import json     # ✅ encode payload SSE
import logging  # ✅ ghi log lỗi/exception
import time     # ✅ minute bucket + UTC timestamp
from pathlib import Path  # ✅ resolve đường dẫn file HTML tĩnh
from typing import Any, Dict, List, Optional, Tuple  # ✅ type hints rõ ràng

# ===== FastAPI =====
from fastapi import APIRouter, Cookie, Depends, HTTPException, Query, Request, status  # ✅ core API
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse  # ✅ response types

# ===== Project imports (giữ theo dự án của bạn) =====
from auth.oauth2 import get_info_user_via_token, required_token_user  # ✅ dùng sẵn của bạn
from security.keyspace import k_metric_5xx, k_metric_bans, k_metric_req  # ✅ key builder cho redis
from security.redis_client import get_redis  # ✅ redis client (sync)
from db.database import get_db

# ===== Third-party =====
from openpyxl import Workbook  # ✅ tạo file xlsx

# =============================================================================
# (0) Router + logger + redis client
# =============================================================================

# Tạo router với prefix /ops để gom endpoint vào 1 namespace.
router = APIRouter(prefix="/ops", tags=["Ops Dashboard"])

# Logger cho module này (để bạn xem lỗi trên stdout hoặc file log tuỳ cấu hình).
logger = logging.getLogger(__name__)

# Lấy redis client dùng chung.
# LƯU Ý: get_redis() trong dự án bạn đang trả về client kiểu sync (redis-py).
# Trong các endpoint sync thì OK, nhưng trong SSE async sẽ offload sang thread.
_r = get_redis()

# =============================================================================
# (1) Config / constants (đặt tập trung để dễ chỉnh)
# =============================================================================

# Khoảng thời gian đẩy snapshot SSE (giây).
SSE_INTERVAL_SECONDS: float = 2.5

# Mặc định thống kê timeseries cho 10 phút gần nhất (tức 10 bucket phút).
DEFAULT_WINDOW_MINUTES: int = 10

# Giới hạn quét scan_iter mỗi vòng (tránh scan quá lớn).
REDIS_SCAN_COUNT: int = 2000

# Giới hạn số item suspicious trả về trong SSE (UI không cần quá nhiều).
SSE_SUSPICIOUS_LIMIT: int = 50

# Mỗi bao nhiêu giây thì SSE mới quét lại danh sách suspicious/bans (tránh quét liên tục).
# Vì scan_iter + ttl rất tốn tài nguyên nếu làm 2.5s/lần.
SSE_SCAN_REFRESH_SECONDS: float = 10.0

# Quyền được phép xem dashboard ops
OPS_ALLOWED_PRIVILEGES = {"Admin", "Boss"}

# =============================================================================
# (2) Helpers: privilege guard, token extraction, redis snapshot
# =============================================================================

def _get_privilege(user: Any) -> Optional[str]:
    """
    Lấy trường Privilege từ user object.
    Hỗ trợ cả kiểu object (có attribute) và dict.
    """
    if isinstance(user, dict):
        return user.get("Privilege")
    return getattr(user, "Privilege", None)


def require_ops_admin(user: Any = Depends(required_token_user)) -> Any:
    """
    Dependency guard cho các endpoint JSON/Export (non-SSE) khi client gửi Authorization header.

    - required_token_user: dependency của bạn (đã verify token).
    - Sau khi có user, kiểm tra Privilege ∈ {Admin, Boss}.

    Trả về user để endpoint có thể dùng thêm (nếu cần).
    """
    priv = _get_privilege(user)
    if priv not in OPS_ALLOWED_PRIVILEGES:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient privilege",
        )
    return user


def _extract_token_from_request(
    request: Request,
    token_from_query: str,
    access_token_cookie: Optional[str],
) -> str:
    """
    Rút token từ 3 nguồn theo thứ tự ưu tiên:
    1) Query param ?token=...
    2) Header Authorization: Bearer <token>
    3) Cookie access_token (HttpOnly)

    Trả về token thô (không bao gồm chữ Bearer) hoặc raise 401 nếu không có.
    """
    # (1) query token (?token=...)
    if token_from_query and token_from_query.strip():
        return token_from_query.strip()

    # (2) Authorization header
    authz = request.headers.get("authorization") or request.headers.get("Authorization") or ""
    authz = authz.strip()
    if authz.lower().startswith("bearer "):
        return authz.split(" ", 1)[1].strip()

    # (3) Cookie token
    if access_token_cookie and access_token_cookie.strip():
        return access_token_cookie.strip()

    # Không có token ở bất kỳ nguồn nào → 401
    raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing token")


def _normalize_bearer(raw_token: str) -> str:
    """
    Chuẩn hoá token thành chuỗi dạng "Bearer <token>" để phù hợp với
    get_info_user_via_token() (nhiều dự án của bạn yêu cầu prefix này).
    """
    t = (raw_token or "").strip()
    if not t:
        # Trường hợp này đã bị chặn từ _extract_token_from_request(), nhưng thêm check để an toàn.
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing token")
    return t if t.lower().startswith("bearer ") else f"Bearer {t}"


def _assert_ops_privilege(user: Any) -> None:
    """
    Kiểm tra quyền Ops (Admin/Boss). Nếu không đạt → raise 403.
    """
    priv = _get_privilege(user)
    if priv not in OPS_ALLOWED_PRIVILEGES:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient privilege")


def _minute_buckets(window_minutes: int) -> List[int]:
    """
    Tạo danh sách minute-bucket cho window N phút gần nhất.
    Ví dụ window=10 → 10 bucket: now-9 ... now
    """
    # Lấy bucket phút hiện tại theo epoch minute.
    now_min = int(time.time() // 60)
    # Trả về list bucket liên tiếp.
    return list(range(now_min - (window_minutes - 1), now_min + 1))


def _redis_get_timeseries(minute_buckets: List[int]) -> Dict[str, List[int]]:
    """
    Đọc 3 timeseries từ Redis theo bucket phút:
    - Requests/min
    - 5xx/min
    - Bans/min

    LƯU Ý: _r.mget là sync, nên khi gọi từ SSE async phải offload bằng asyncio.to_thread().
    """
    # Tạo key list theo từng metric.
    keys_req = [k_metric_req(m) for m in minute_buckets]
    keys_5xx = [k_metric_5xx(m) for m in minute_buckets]
    keys_ban = [k_metric_bans(m) for m in minute_buckets]

    # Redis mget trả về list bytes/str/None tuỳ cấu hình decode_responses.
    v_req = _r.mget(keys_req)
    v_5xx = _r.mget(keys_5xx)
    v_ban = _r.mget(keys_ban)

    # Hàm convert an toàn: None → 0, bytes/str → int
    def to_int(v: Any) -> int:
        try:
            if v is None:
                return 0
            # redis-py có thể trả bytes; int(b'123') sẽ lỗi, nên decode trước.
            if isinstance(v, (bytes, bytearray)):
                v = v.decode("utf-8", "ignore")
            return int(v)
        except Exception:
            return 0

    return {
        "req": [to_int(x) for x in v_req],
        "s5xx": [to_int(x) for x in v_5xx],
        "bans": [to_int(x) for x in v_ban],
    }


def _scan_suspicious_top(limit: int) -> List[Dict[str, Any]]:
    """
    Quét các key dạng sus:ip:*:5min và trả danh sách top nghi vấn.

    Tối ưu:
    - scan_iter để lấy danh sách key (có giới hạn count)
    - pipeline GET + TTL để giảm round-trip
    """
    # Thu thập key suspicious (giới hạn số key xử lý để tránh quá tải).
    keys: List[bytes] = []
    for k in _r.scan_iter(match=b"sus:ip:*:5min", count=REDIS_SCAN_COUNT):
        keys.append(k)
        # Hard cap để bảo vệ server nếu key quá nhiều.
        if len(keys) >= 5000:
            break

    if not keys:
        return []

    # Pipeline GET + TTL theo cặp (get, ttl) cho từng key.
    pipe = _r.pipeline()
    for k in keys:
        pipe.get(k)
        pipe.ttl(k)
    results = pipe.execute()

    # Parse kết quả.
    out: List[Dict[str, Any]] = []
    prefix, suffix = "sus:ip:", ":5min"
    for i, k in enumerate(keys):
        # Mỗi key có 2 result: get và ttl.
        v_score = results[2 * i]
        v_ttl = results[2 * i + 1]

        # Decode key để lấy IP
        k_str = k.decode("utf-8", "ignore")
        if not (k_str.startswith(prefix) and k_str.endswith(suffix)):
            continue
        ip = k_str[len(prefix) : -len(suffix)]

        # Convert score/ttl an toàn
        try:
            if isinstance(v_score, (bytes, bytearray)):
                v_score = v_score.decode("utf-8", "ignore")
            score = int(v_score or 0)
        except Exception:
            score = 0

        try:
            ttl = int(v_ttl) if v_ttl is not None else -1
        except Exception:
            ttl = -1

        out.append(
            {
                "ip": ip,
                "score": score,
                # ttl<=0 nghĩa là không tồn tại hoặc không có expiry, ta trả None để UI hiển thị "-"
                "ttl_seconds": ttl if ttl and ttl > 0 else None,
            }
        )

    # Sort theo score giảm dần, tie-break theo ttl (ttl lớn hơn ưu tiên hơn tuỳ bạn).
    out.sort(key=lambda x: (x["score"], x["ttl_seconds"] or 0), reverse=True)
    return out[: max(1, limit)]


def _scan_current_bans() -> List[Dict[str, Any]]:
    """
    Quét các key dạng ban:ip:* và trả danh sách đang ban (có TTL > 0).
    Pipeline TTL để giảm round-trip.
    """
    keys: List[bytes] = []
    for k in _r.scan_iter(match=b"ban:ip:*", count=REDIS_SCAN_COUNT):
        keys.append(k)
        if len(keys) >= 5000:
            break

    if not keys:
        return []

    # Pipeline TTL theo từng key
    pipe = _r.pipeline()
    for k in keys:
        pipe.ttl(k)
    ttls = pipe.execute()

    out: List[Dict[str, Any]] = []
    prefix = "ban:ip:"
    for k, ttl in zip(keys, ttls):
        k_str = k.decode("utf-8", "ignore")
        if not k_str.startswith(prefix):
            continue
        ip = k_str[len(prefix) :]

        try:
            ttl_int = int(ttl) if ttl is not None else -1
        except Exception:
            ttl_int = -1

        # Chỉ lấy bans còn TTL
        if ttl_int and ttl_int > 0:
            out.append({"ip": ip, "ttl_seconds": ttl_int})

    # TTL tăng dần (ai sắp hết hạn hiển thị trước)
    out.sort(key=lambda x: x["ttl_seconds"])
    return out


def _build_snapshot_payload(
    window_minutes: int,
    suspicious_limit: int,
    include_scans: bool,
) -> Dict[str, Any]:
    """
    Tạo payload đầy đủ cho SSE:
    - timeseries window_minutes gần nhất
    - (tuỳ chọn) quét suspicious + current bans

    include_scans=False để tái sử dụng dữ liệu quét trước đó (giảm tải).
    """
    minutes = _minute_buckets(window_minutes)
    series = _redis_get_timeseries(minutes)

    payload: Dict[str, Any] = {
        "minutes": minutes,
        "req": series["req"],
        "s5xx": series["s5xx"],
        "bans": series["bans"],
    }

    if include_scans:
        susp_items = _scan_suspicious_top(suspicious_limit)
        ban_items = _scan_current_bans()
        payload["suspicious"] = {"count": len(susp_items), "items": susp_items}
        payload["current_bans"] = {"items": ban_items}

    return payload


# =============================================================================
# (3) (A) HTML dashboard: trả file tĩnh
# =============================================================================

@router.get("/dashboard", response_class=HTMLResponse, summary="Dashboard HTML tách file")
def ops_dashboard_html() -> FileResponse:
    """
    Trả file HTML tĩnh.

    LƯU Ý:
    - Path(__file__).resolve().parents[1] giả định cấu trúc: app/routers/ops_dashboard.py
      và app/static/ops_dashboard.html.
    - Nếu bạn đổi cấu trúc, hãy điều chỉnh parents[...] cho đúng.
    """
    html_path = Path(__file__).resolve().parents[1] / "static" / "ops_dashboard.html"
    if not html_path.exists():
        raise HTTPException(status_code=500, detail=f"Không tìm thấy file: {html_path}")
    return FileResponse(str(html_path), media_type="text/html; charset=utf-8")


# =============================================================================
# (4) (B) SSE: realtime metrics
# =============================================================================

@router.get("/sse", summary="Server-Sent Events: realtime metrics")
async def sse_metrics(
    request: Request,
    token: str = Query("", alias="token"),
    access_token: Optional[str] = Cookie(default=None),
):
    # 1) Lấy token thô từ request (query → header → cookie)
    raw_token = _extract_token_from_request(request, token, access_token)

    # 2) Chuẩn hoá token dạng "Bearer ..." (phục vụ nhất quán)
    normalized = _normalize_bearer(raw_token)

    # 3) Vì get_info_user_via_token() đang dùng Depends(get_db),
    #    nên nếu bạn gọi trực tiếp như function thì FastAPI KHÔNG inject db.
    #    Ta sẽ tự tạo DB session và truyền thẳng vào tham số db=...
    #
    #    Đồng thời: jwt.decode trong get_info_user_via_token() thường cần RAW JWT,
    #    không cần "Bearer ". Vì vậy ta strip "Bearer " nếu có.
    jwt_token = normalized.split(" ", 1)[1].strip()  # bỏ "Bearer "

    # 4) Tạo helper chạy trong thread:
    #    - tạo session trong chính thread đó (Session không thread-safe)
    #    - gọi get_info_user_via_token(token=..., db=...)
    #    - đảm bảo đóng session
    def _verify_user_in_thread(token_str: str):
        db_gen = get_db()  # generator yield Session
        db = next(db_gen)  # lấy Session thật

        try:
            # ✅ gọi function giữ nguyên cấu trúc, nhưng truyền token & db thật
            return get_info_user_via_token(token=token_str, db=db)
        finally:
            # đóng generator để trigger finally trong get_db() (đóng session)
            db_gen.close()

    # 5) Verify token + lấy user (chạy trong thread để không block event-loop)
    try:
        user = await asyncio.to_thread(_verify_user_in_thread, jwt_token)
    except HTTPException as ex:
        # nếu get_info_user_via_token raise HTTPException (401) thì giữ nguyên
        logger.warning("SSE auth failed: %s", ex)
        raise
    except StopIteration:
        # trường hợp get_db() không yield được session (hiếm, nhưng nên bắt)
        logger.exception("SSE auth failed: get_db() did not yield a session")
        raise HTTPException(status_code=500, detail="Database session unavailable")
    except Exception as ex:
        logger.warning("SSE auth failed: %s", ex)
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid or expired token")

    # 6) Nếu token không có hoặc decode fail mà bạn return None, chặn luôn:
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid or expired token")

    # 7) Kiểm quyền ops
    _assert_ops_privilege(user)

    # 8) Generator SSE
    async def _gen():
        yield "retry: 2500\n\n"

        last_scan_ts = 0.0
        cached_susp: Dict[str, Any] = {"count": 0, "items": []}
        cached_bans: Dict[str, Any] = {"items": []}

        while True:
            if await request.is_disconnected():
                break

            try:
                now = time.time()
                need_scan = (now - last_scan_ts) >= SSE_SCAN_REFRESH_SECONDS

                payload = await asyncio.to_thread(
                    _build_snapshot_payload,
                    DEFAULT_WINDOW_MINUTES,
                    SSE_SUSPICIOUS_LIMIT,
                    need_scan,
                )

                if need_scan:
                    cached_susp = payload.get("suspicious", cached_susp)
                    cached_bans = payload.get("current_bans", cached_bans)
                    last_scan_ts = now
                else:
                    payload["suspicious"] = cached_susp
                    payload["current_bans"] = cached_bans

                data = json.dumps(payload, ensure_ascii=False)
                yield f"event: metrics\ndata: {data}\n\n"

            except asyncio.CancelledError:
                break
            except Exception as ex:
                logger.exception("SSE loop error: %s", ex)
                yield ":keepalive\n\n"

            await asyncio.sleep(SSE_INTERVAL_SECONDS)

    headers = {
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "X-Accel-Buffering": "no",
    }

    return StreamingResponse(_gen(), media_type="text/event-stream; charset=utf-8", headers=headers)


# =============================================================================
# (5) (C) API: Summary timeseries (10 phút gần nhất) – dùng fallback/Excel
# =============================================================================

@router.get("/metrics/summary", summary="Timeseries req/min, 5xx/min, bans/min (10 phút gần nhất)")
def metrics_summary(_: Any = Depends(require_ops_admin)):
    """
    Endpoint JSON trả timeseries 10 phút gần nhất.
    - Dùng cho fallback nếu SSE có vấn đề.
    - Dùng để kiểm tra nhanh server/redis.
    """
    minutes = _minute_buckets(DEFAULT_WINDOW_MINUTES)
    series = _redis_get_timeseries(minutes)
    return {"minutes": minutes, **series}


# =============================================================================
# (6) (D) API: Top-N suspicious
# =============================================================================

@router.get("/metrics/top_suspicious", summary="Top-N IP nghi vấn còn TTL (quét sus:ip:*:5min)")
def top_suspicious(limit: int = 50, _: Any = Depends(require_ops_admin)):
    """
    Trả Top-N IP nghi vấn.
    - limit mặc định 50.
    - Quét trực tiếp redis (sync endpoint → FastAPI chạy trong threadpool).
    """
    if limit < 1 or limit > 500:
        raise HTTPException(status_code=400, detail="limit phải trong [1, 500]")
    items = _scan_suspicious_top(limit)
    return {"count": len(items), "items": items}


# =============================================================================
# (7) (E) API: Current bans
# =============================================================================

@router.get("/metrics/current_bans", summary="Danh sách IP đang bị BAN + TTL")
def current_bans(_: Any = Depends(require_ops_admin)):
    """
    Trả danh sách ip bị ban còn TTL.
    """
    items = _scan_current_bans()
    return {"items": items}


# =============================================================================
# (8) (F) Export Excel
# =============================================================================

@router.get("/export/metrics.xlsx", summary="Xuất Excel tổng hợp (summary + suspicious + bans)")
def export_excel(minutes: int = DEFAULT_WINDOW_MINUTES, _: Any = Depends(require_ops_admin)):
    """
    Xuất file Excel gồm 3 sheet:
    - Summary (timeseries theo minute bucket)
    - TopSuspicious
    - CurrentBans

    minutes: số phút cần xuất (1..240). Mặc định 10.
    """
    # Validate input
    if minutes < 1 or minutes > 240:
        raise HTTPException(status_code=400, detail="minutes phải trong [1, 240]")

    # 1) Timeseries theo bucket phút
    mins = _minute_buckets(minutes)
    series = _redis_get_timeseries(mins)

    # 2) Quét suspicious/bans
    susp_items = _scan_suspicious_top(limit=500)  # Excel cho phép nhiều hơn SSE
    ban_items = _scan_current_bans()

    # 3) Tạo workbook
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["MinuteBucket", "Timestamp(UTC)", "Requests", "5xx", "Bans"])
    for m, req, s5xx, bans in zip(mins, series["req"], series["s5xx"], series["bans"]):
        iso_utc = time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime(m * 60))
        ws1.append([m, iso_utc, req, s5xx, bans])

    # --- Sheet 2: TopSuspicious ---
    ws2 = wb.create_sheet("TopSuspicious")
    ws2.append(["IP", "Score(5min)", "TTL(s)"])
    for it in susp_items:
        ws2.append([it.get("ip"), it.get("score"), it.get("ttl_seconds")])

    # --- Sheet 3: CurrentBans ---
    ws3 = wb.create_sheet("CurrentBans")
    ws3.append(["IP", "TTL(s)"])
    for it in ban_items:
        ws3.append([it.get("ip"), it.get("ttl_seconds")])

    # 4) Ghi workbook vào memory (BytesIO) để trả về StreamingResponse
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    # 5) Headers tải file
    headers = {
        "Content-Disposition": 'attachment; filename="ops_metrics.xlsx"',
        "Cache-Control": "no-store",
    }

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
